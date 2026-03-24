"""
Budget Report Generator
Läser in exporterad .xlsx-data från Budgetplaneraren och skapar
en snygg, färgkodad Excel-rapport med diagram.

Användning:
    python budget_report.py budget_tracker.xlsx
    python budget_report.py budget_tracker.xlsx -o rapport.xlsx
"""

import sys
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installera openpyxl: pip install openpyxl")
    sys.exit(1)


# ── Färgschema (matchar HTML-sektionerna) ────────────────────
SECTION_COLORS = {
    "Inkomster":        {"color": "10B981", "light": "D1FAE5", "bg": "ECFDF5"},
    "Bostadskostnader": {"color": "3B6DE0", "light": "DBEAFE", "bg": "EFF6FF"},
    "Försäkringar":     {"color": "E85D5D", "light": "F5C2C2", "bg": "FDEAEA"},
    "Prenumerationer":  {"color": "FF9D4D", "light": "FED7AA", "bg": "FFF7ED"},
    "Sparande":         {"color": "005706", "light": "D1FAE5", "bg": "ECFDF5"},
    "Övrigt":           {"color": "FBBF24", "light": "FEE2B3", "bg": "FFF7ED"},
}
DEFAULT_COLOR = {"color": "6B7280", "light": "E5E7EB", "bg": "F9FAFB"}

PIE_FILLS = ["3B6DE0", "E85D5D", "FF9D4D", "005706", "FBBF24",
             "8B5CF6", "DB2777", "059669", "DC2626", "0891B2"]


def get_colors(name):
    return SECTION_COLORS.get(name, DEFAULT_COLOR)


def parse_input(filepath):
    """Läser den exporterade xlsx-filen och returnerar strukturerad data."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    summary = {}
    sections = []
    current_section = None
    mode = "summary"

    i = 0
    while i < len(rows):
        row = rows[i]
        cell0 = str(row[0]).strip() if row[0] else ""

        # Sammanfattningsrader
        if cell0 == "Totala inkomster" and row[1] is not None:
            summary["income"] = float(row[1])
        elif cell0 == "Totalt budgeterat" and row[1] is not None:
            summary["budget"] = float(row[1])
        elif cell0 == "Totalt utgifter" and row[1] is not None:
            summary["spent"] = float(row[1])
        elif cell0 == "Återstår" and row[1] is not None:
            summary["remaining"] = float(row[1])

        # Sektionsrubriker i detaljdelen
        if cell0 == "DETAILED BREAKDOWN":
            mode = "detail"
            i += 1
            continue

        if mode == "detail" and cell0:
            # Kolla om det är en sektionsrubrik (VERSALER, enbart text i kolumn A)
            is_header = (row[1] is None or str(row[1]).strip() == "") and cell0 == cell0.upper() and len(cell0) > 2
            # Kolla om det är en underrubrik (Post, Budget, ...)
            is_subheader = cell0 in ("Post",) and row[1] is not None

            if is_header:
                name = cell0.title()
                current_section = {"name": name, "items": [], "total_budget": 0, "total_actual": 0}
                sections.append(current_section)
            elif is_subheader:
                pass  # hoppa över rubrikrad
            elif current_section and row[1] is not None:
                try:
                    budget = float(row[1]) if row[1] else 0
                    actual = float(row[2]) if row[2] else 0
                    current_section["items"].append({
                        "name": str(row[0]),
                        "budget": budget,
                        "actual": actual,
                    })
                    current_section["total_budget"] += budget
                    current_section["total_actual"] += actual
                except (ValueError, TypeError):
                    pass

        i += 1

    return summary, sections


def build_report(summary, sections, output_path):
    """Bygger den snygga Excel-rapporten."""
    wb = openpyxl.Workbook()

    # ── Typsnitt & stilar ────────────────────────────────────
    title_font = Font(name="Calibri", size=18, bold=True, color="1E293B")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    item_font = Font(name="Calibri", size=11, color="333333")
    value_font = Font(name="Calibri", size=11, color="333333")
    pct_font_ok = Font(name="Calibri", size=11, color="10B981", bold=True)
    pct_font_warn = Font(name="Calibri", size=11, color="F59E0B", bold=True)
    pct_font_over = Font(name="Calibri", size=11, color="DC2626", bold=True)
    kpi_label_font = Font(name="Calibri", size=10, bold=True, color="6B7280")
    kpi_value_font = Font(name="Calibri", size=16, bold=True, color="1E293B")

    thin_border_side = Side(style="thin", color="E5E7EB")
    cell_border = Border(
        bottom=thin_border_side,
    )

    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ────────────────────────────────────────────────────────
    #  SAMMANFATTNING
    # ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sammanfattning"
    ws.sheet_properties.tabColor = "1E293B"

    # Kolumnbredder
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    # Titel
    ws.merge_cells("B2:F2")
    ws["B2"] = "BUDGETPLANERARE — RAPPORT"
    ws["B2"].font = title_font

    # KPI-kort
    kpis = [
        ("Totala inkomster", summary.get("income", 0), "10B981"),
        ("Totalt budgeterat", summary.get("budget", 0), "3B6DE0"),
        ("Totala utgifter", summary.get("spent", 0), "DC2626"),
        ("Återstår", summary.get("remaining", 0), "2563EB"),
    ]
    row = 4
    for col_idx, (label, value, color) in enumerate(kpis):
        c = col_idx + 2  # B=2, C=3, ...
        cell_label = ws.cell(row=row, column=c, value=label)
        cell_label.font = Font(name="Calibri", size=9, bold=True, color=color)
        cell_label.alignment = center
        cell_value = ws.cell(row=row + 1, column=c, value=f"{value:,.0f} kr".replace(",", " "))
        cell_value.font = Font(name="Calibri", size=14, bold=True, color=color)
        cell_value.alignment = center

    # ── Sektionsöversikt ─────────────────────────────────────
    row = 8
    headers = ["Avsnitt", "Budgeterat", "Utgifter", "Återstår", "% Använt"]
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    for col_idx, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_idx + 2, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center if col_idx > 0 else left_align

    row += 1
    expense_sections = [s for s in sections if s["name"] != "Inkomster"]
    for s in sections:
        colors = get_colors(s["name"])
        bg_fill = PatternFill(start_color=colors["bg"].replace("#", ""),
                              end_color=colors["bg"].replace("#", ""), fill_type="solid")
        remaining = s["total_budget"] - s["total_actual"]
        pct_val = round((s["total_actual"] / s["total_budget"]) * 100) if s["total_budget"] > 0 else 0

        ws.cell(row=row, column=2, value=s["name"]).font = Font(
            name="Calibri", size=11, bold=True, color=colors["color"])
        ws.cell(row=row, column=3, value=s["total_budget"]).number_format = '#,##0 "kr"'
        ws.cell(row=row, column=4, value=s["total_actual"]).number_format = '#,##0 "kr"'
        ws.cell(row=row, column=5, value=remaining).number_format = '#,##0 "kr"'
        pct_cell = ws.cell(row=row, column=6, value=f"{pct_val}%")
        if pct_val > 100:
            pct_cell.font = pct_font_over
        elif pct_val > 80:
            pct_cell.font = pct_font_warn
        else:
            pct_cell.font = pct_font_ok
        pct_cell.alignment = center

        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = bg_fill
            ws.cell(row=row, column=c).border = cell_border
            if c in (3, 4, 5):
                ws.cell(row=row, column=c).alignment = right_align

        row += 1

    # ── Cirkeldiagram (Utgiftsfördelning) ────────────────────
    pie_data_start = row + 2
    ws.cell(row=pie_data_start, column=2, value="Avsnitt").font = kpi_label_font
    ws.cell(row=pie_data_start, column=3, value="Utgifter").font = kpi_label_font
    for idx, s in enumerate(expense_sections):
        ws.cell(row=pie_data_start + 1 + idx, column=2, value=s["name"])
        ws.cell(row=pie_data_start + 1 + idx, column=3, value=s["total_actual"])
    pie_end = pie_data_start + len(expense_sections)

    pie = PieChart()
    pie.title = "Utgiftsfördelning"
    pie.style = 10
    pie.width = 16
    pie.height = 12
    data_ref = Reference(ws, min_col=3, min_row=pie_data_start,
                         max_row=pie_end)
    cats_ref = Reference(ws, min_col=2, min_row=pie_data_start + 1,
                         max_row=pie_end)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False

    for i, s in enumerate(expense_sections):
        colors = get_colors(s["name"])
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = colors["color"]
        pie.series[0].data_points.append(pt)

    ws.add_chart(pie, f"B{pie_end + 2}")

    # ── Stapeldiagram (Budget vs Utgifter) ───────────────────
    bar_data_start = pie_end + 2
    bar_row = bar_data_start
    ws.cell(row=bar_row, column=5, value="Avsnitt").font = kpi_label_font
    ws.cell(row=bar_row, column=6, value="Budget").font = kpi_label_font
    ws.cell(row=bar_row, column=7, value="Utgifter").font = kpi_label_font
    ws.column_dimensions["G"].width = 14
    for idx, s in enumerate(expense_sections):
        ws.cell(row=bar_row + 1 + idx, column=5, value=s["name"])
        ws.cell(row=bar_row + 1 + idx, column=6, value=s["total_budget"])
        ws.cell(row=bar_row + 1 + idx, column=7, value=s["total_actual"])
    bar_end = bar_row + len(expense_sections)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Budget kontra utgifter"
    bar.style = 10
    bar.width = 16
    bar.height = 12
    bar.y_axis.numFmt = '#,##0'

    budget_ref = Reference(ws, min_col=6, min_row=bar_row,
                           max_row=bar_end)
    actual_ref = Reference(ws, min_col=7, min_row=bar_row,
                           max_row=bar_end)
    cats_bar = Reference(ws, min_col=5, min_row=bar_row + 1,
                         max_row=bar_end)
    bar.add_data(budget_ref, titles_from_data=True)
    bar.add_data(actual_ref, titles_from_data=True)
    bar.set_categories(cats_bar)

    bar.series[0].graphicalProperties.solidFill = "DBEAFE"
    bar.series[0].graphicalProperties.line.solidFill = "3B6DE0"
    bar.series[1].graphicalProperties.solidFill = "3B6DE0"
    bar.series[1].graphicalProperties.line.solidFill = "1E40AF"

    ws.add_chart(bar, f"E{bar_end + 2}")

    # ────────────────────────────────────────────────────────
    #  DETALJBLAD
    # ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detaljerad uppdelning")
    ws2.sheet_properties.tabColor = "3B6DE0"

    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 12

    ws2.merge_cells("B2:E2")
    ws2["B2"] = "DETALJERAD UPPDELNING"
    ws2["B2"].font = title_font

    row = 4
    for s in sections:
        colors = get_colors(s["name"])
        section_fill = PatternFill(start_color=colors["color"],
                                   end_color=colors["color"], fill_type="solid")
        bg_fill = PatternFill(start_color=colors["bg"].replace("#", ""),
                              end_color=colors["bg"].replace("#", ""), fill_type="solid")

        # Sektionsrubrik
        for c in range(2, 6):
            ws2.cell(row=row, column=c).fill = section_fill
        ws2.cell(row=row, column=2, value=s["name"]).font = section_font
        is_income = s["name"] == "Inkomster"
        ws2.cell(row=row, column=3, value="Förväntat" if is_income else "Budget").font = header_font
        ws2.cell(row=row, column=3).alignment = right_align
        ws2.cell(row=row, column=4, value="Mottaget" if is_income else "Utgifter").font = header_font
        ws2.cell(row=row, column=4).alignment = right_align
        if not is_income:
            ws2.cell(row=row, column=5, value="%").font = header_font
            ws2.cell(row=row, column=5).alignment = center
        row += 1

        for item in s["items"]:
            pct_val = round((item["actual"] / item["budget"]) * 100) if item["budget"] > 0 else 0

            ws2.cell(row=row, column=2, value=item["name"]).font = item_font
            ws2.cell(row=row, column=3, value=item["budget"]).font = value_font
            ws2.cell(row=row, column=3).number_format = '#,##0'
            ws2.cell(row=row, column=3).alignment = right_align
            ws2.cell(row=row, column=4, value=item["actual"]).font = value_font
            ws2.cell(row=row, column=4).number_format = '#,##0'
            ws2.cell(row=row, column=4).alignment = right_align

            if not is_income:
                pct_cell = ws2.cell(row=row, column=5, value=f"{pct_val}%")
                if pct_val > 100:
                    pct_cell.font = pct_font_over
                elif pct_val > 80:
                    pct_cell.font = pct_font_warn
                else:
                    pct_cell.font = pct_font_ok
                pct_cell.alignment = center

            for c in range(2, 6):
                ws2.cell(row=row, column=c).fill = bg_fill
                ws2.cell(row=row, column=c).border = cell_border

            row += 1

        # Summeringsrad
        sum_fill = PatternFill(start_color=colors["light"].replace("#", ""),
                               end_color=colors["light"].replace("#", ""), fill_type="solid")
        total_pct = round((s["total_actual"] / s["total_budget"]) * 100) if s["total_budget"] > 0 else 0

        ws2.cell(row=row, column=2, value="Totalt").font = Font(
            name="Calibri", size=11, bold=True, color=colors["color"])
        ws2.cell(row=row, column=3, value=s["total_budget"]).font = Font(
            name="Calibri", size=11, bold=True, color="333333")
        ws2.cell(row=row, column=3).number_format = '#,##0'
        ws2.cell(row=row, column=3).alignment = right_align
        ws2.cell(row=row, column=4, value=s["total_actual"]).font = Font(
            name="Calibri", size=11, bold=True, color="333333")
        ws2.cell(row=row, column=4).number_format = '#,##0'
        ws2.cell(row=row, column=4).alignment = right_align

        if not is_income:
            tot_cell = ws2.cell(row=row, column=5, value=f"{total_pct}%")
            tot_cell.font = Font(name="Calibri", size=11, bold=True,
                                 color="DC2626" if total_pct > 100
                                 else "F59E0B" if total_pct > 80
                                 else "10B981")
            tot_cell.alignment = center

        for c in range(2, 6):
            ws2.cell(row=row, column=c).fill = sum_fill
            ws2.cell(row=row, column=c).border = Border(
                top=Side(style="medium", color=colors["color"]),
                bottom=Side(style="medium", color=colors["color"]),
            )

        row += 2  # mellanrum

    wb.save(output_path)
    print(f"✓ Rapport sparad: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generera budgetrapport från exporterad xlsx")
    parser.add_argument("input", help="Sökväg till exporterad budget_tracker.xlsx")
    parser.add_argument("-o", "--output", default=None,
                        help="Sökväg för utdata (standard: budget_rapport.xlsx)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Filen hittades inte: {input_path}")
        sys.exit(1)

    output_path = args.output or input_path.parent / "budget_rapport.xlsx"

    summary, sections = parse_input(input_path)
    build_report(summary, sections, str(output_path))


if __name__ == "__main__":
    main()
