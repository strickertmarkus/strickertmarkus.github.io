"""
Professional Budget Tracker – Google Sheets Compatible (.xlsx)
Generates a multi-sheet workbook with:
  • Dashboard   – donut charts per category showing spend-vs-budget
  • Monthly Log – single entry point for all transactions
  • Category sheets with formulas that auto-aggregate from the log
  • Colour-coded sections, conditional formatting, and summary stats
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import DoughnutChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
from openpyxl.utils import get_column_letter
from copy import copy

# ── Colour palette ──────────────────────────────────────────────────
COLORS = {
    "Income":        {"primary": "2E7D32", "light": "C8E6C9", "accent": "1B5E20"},
    "House Loan":    {"primary": "1565C0", "light": "BBDEFB", "accent": "0D47A1"},
    "Insurances":    {"primary": "6A1B9A", "light": "E1BEE7", "accent": "4A148C"},
    "Subscriptions": {"primary": "E65100", "light": "FFE0B2", "accent": "BF360C"},
    "Savings":       {"primary": "00838F", "light": "B2EBF2", "accent": "006064"},
    "Misc.":         {"primary": "455A64", "light": "CFD8DC", "accent": "263238"},
}

CATEGORY_ORDER = ["Income", "House Loan", "Insurances", "Subscriptions", "Savings", "Misc."]

# Example budget items per category (user can add rows freely)
BUDGET_ITEMS = {
    "Income": [
        ("Salary (net)", 35000),
        ("Side income", 0),
        ("Other income", 0),
    ],
    "House Loan": [
        ("Mortgage payment", 8500),
        ("Amortisation", 3000),
        ("HOA / Bostadsrättsavgift", 4200),
        ("Home insurance (building)", 350),
        ("Electricity", 800),
        ("Internet", 399),
    ],
    "Insurances": [
        ("Health insurance", 450),
        ("Car insurance", 600),
        ("Life insurance", 300),
        ("Pet insurance", 250),
    ],
    "Subscriptions": [
        ("Spotify", 129),
        ("Netflix", 169),
        ("Cloud storage", 99),
        ("Gym membership", 399),
        ("News / magazine", 149),
    ],
    "Savings": [
        ("Emergency fund", 2000),
        ("Investment account", 3000),
        ("Vacation fund", 1000),
        ("Retirement (extra)", 1500),
    ],
    "Misc.": [
        ("Groceries", 5000),
        ("Going out meals", 1000),
        ("Transportation", 1200),
        ("Clothing", 500),
        ("Entertainment", 500),
        ("Personal care", 300),
        ("Gifts", 300),
    ],
}

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# ── Style helpers ───────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

HEADER_BORDER = Border(
    left=Side(style="thin", color="757575"),
    right=Side(style="thin", color="757575"),
    top=Side(style="thin", color="757575"),
    bottom=Side(style="medium", color="424242"),
)


def style_cell(ws, row, col, value=None, font=None, fill=None, alignment=None,
               border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(vertical="center")
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def fill_range(ws, row_start, row_end, col_start, col_end, fill, border=None):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            if border:
                cell.border = border


def sek_fmt(value_str):
    """Return a SEK number-format string for openpyxl."""
    return '#,##0 "kr"'


# ── 1. DASHBOARD SHEET ─────────────────────────────────────────────
def build_dashboard(wb, summary_sheet_name="Summary"):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = "37474F"

    # Column widths
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Title
    ws.merge_cells("A1:R1")
    style_cell(ws, 1, 1, "BUDGET TRACKER — DASHBOARD",
               font=Font(name="Calibri", size=22, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 52

    ws.merge_cells("A2:R2")
    style_cell(ws, 2, 1, 'Data auto-populated from the "Summary" sheet. Register transactions in the Monthly Log.',
               font=Font(name="Calibri", size=10, italic=True, color="90A4AE"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 24
    fill_range(ws, 2, 2, 1, 18, PatternFill("solid", fgColor="37474F"))

    # ── Top summary bar row 4-6 ──
    fill_range(ws, 3, 3, 1, 18, PatternFill("solid", fgColor="ECEFF1"))
    labels = ["Total Income", "Total Budgeted", "Total Spent", "Remaining"]
    cols_start = [1, 5, 9, 13]
    for i, (lbl, cs) in enumerate(zip(labels, cols_start)):
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=cs + 3)
        style_cell(ws, 4, cs, lbl,
                   font=Font(name="Calibri", size=11, bold=True, color="546E7A"),
                   fill=PatternFill("solid", fgColor="ECEFF1"),
                   alignment=Alignment(horizontal="center", vertical="center"))
        fill_range(ws, 4, 4, cs, cs + 3, PatternFill("solid", fgColor="ECEFF1"))

        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=cs + 3)
        # Formulas referencing Summary sheet
        formulas = [
            f"=Summary!C3",   # Total Income
            f"=Summary!C5",   # Total Budgeted
            f"=Summary!C7",   # Total Spent
            f"=Summary!C9",   # Remaining
        ]
        style_cell(ws, 5, cs, formulas[i],
                   font=Font(name="Calibri", size=18, bold=True,
                             color="2E7D32" if i == 0 else ("C62828" if i == 2 else "37474F")),
                   fill=PatternFill("solid", fgColor="ECEFF1"),
                   alignment=Alignment(horizontal="center", vertical="center"),
                   number_format=sek_fmt(""))
        fill_range(ws, 5, 5, cs, cs + 3, PatternFill("solid", fgColor="ECEFF1"))

    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 38
    fill_range(ws, 6, 6, 1, 18, PatternFill("solid", fgColor="ECEFF1"))

    # ── Donut charts per category (rows 8+) ──
    chart_row = 8
    chart_col_positions = [1, 10]  # Two chart columns per row

    expense_cats = [c for c in CATEGORY_ORDER if c != "Income"]
    for idx, cat in enumerate(expense_cats):
        c_offset = chart_col_positions[idx % 2]
        
        # Vertical spacing: 2 categories per row, each takes ~25 rows
        row_num = (idx // 2)
        r_offset = row_num * 28
        data_row = chart_row + r_offset

        pal = COLORS[cat]
        items = BUDGET_ITEMS[cat]

        # ── Category header ──
        ws.merge_cells(start_row=data_row, start_column=c_offset,
                       end_row=data_row, end_column=c_offset + 7)
        style_cell(ws, data_row, c_offset, cat.upper(),
                   font=Font(name="Calibri", size=13, bold=True, color=pal["primary"]),
                   fill=PatternFill("solid", fgColor=pal["light"]),
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=THIN_BORDER)
        fill_range(ws, data_row, data_row, c_offset, c_offset + 7,
                   PatternFill("solid", fgColor=pal["light"]), THIN_BORDER)
        ws.row_dimensions[data_row].height = 28

        # ── Build data for multi-segment donut: one segment per item ──
        # Row 1: item names (labels)
        # Row 2: spent amounts (for coloring)
        data_start_row = data_row + 1
        
        # Labels row
        for j, (item_name, _) in enumerate(items):
            col = c_offset + j
            style_cell(ws, data_start_row, col, item_name,
                       font=Font(name="Calibri", size=8, color="424242"))
        
        # Spent amounts row
        for j, (item_name, _) in enumerate(items):
            col = c_offset + j
            cat_sheet_name = cat.replace(".", "")
            item_spent_ref = f"='{cat}'!E{4 + j}"
            style_cell(ws, data_start_row + 1, col, item_spent_ref,
                       number_format=sek_fmt(""))

        # ── Build donut chart with multiple segments ──
        chart = DoughnutChart()
        chart.style = 26
        chart.title = None
        chart.holeSize = 85  # Large hole for thin ring effect
        
        # Data: spent amounts (one segment per item)
        spent_ref = Reference(ws, min_col=c_offset, max_col=c_offset + len(items) - 1,
                              min_row=data_start_row + 1, max_row=data_start_row + 1)
        labels_ref = Reference(ws, min_col=c_offset, max_col=c_offset + len(items) - 1,
                               min_row=data_start_row, max_row=data_start_row)
        
        chart.add_data(spent_ref, from_rows=True)
        chart.set_categories(labels_ref)
        
        series = chart.series[0]
        
        # Color all spent segments with category color
        for j in range(len(items)):
            pt = DataPoint(idx=j)
            pt.graphicalProperties.solidFill = pal["primary"]
            series.data_points.append(pt)

        chart.width = 11
        chart.height = 10
        chart.legend = None

        # Data labels — show item names and percentages
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = True
        dlbl_font = CharacterProperties(
            latin=DrawingFont(typeface="Calibri"), sz=800, solidFill=pal["accent"]
        )
        chart.dataLabels.txPr = RichText(
            p=[Paragraph(pPr=ParagraphProperties(defRPr=dlbl_font), endParaRPr=dlbl_font)]
        )

        # Position chart to the right of the category label, well below headers to avoid overlap
        chart_anchor = f"{get_column_letter(c_offset + 2)}{data_start_row + 5}"
        ws.add_chart(chart, chart_anchor)

        # ── Summary stats below chart ──
        info_row = data_start_row + 14
        
        style_cell(ws, info_row, c_offset, "Budget:",
                   font=Font(name="Calibri", size=10, bold=True, color=pal["accent"]))
        style_cell(ws, info_row, c_offset + 1, f"=Summary!C{_summary_cat_row(cat)}",
                   font=Font(name="Calibri", size=10, color="424242"),
                   number_format=sek_fmt(""))
        
        style_cell(ws, info_row, c_offset + 3, "Spent:",
                   font=Font(name="Calibri", size=10, bold=True, color=pal["accent"]))
        style_cell(ws, info_row, c_offset + 4, f"=Summary!D{_summary_cat_row(cat)}",
                   font=Font(name="Calibri", size=10, color="424242"),
                   number_format=sek_fmt(""))
        
        style_cell(ws, info_row, c_offset + 6, "Used:",
                   font=Font(name="Calibri", size=10, bold=True, color=pal["accent"]))
        pct_formula = (
            f'=IFERROR(Summary!D{_summary_cat_row(cat)}'
            f'/Summary!C{_summary_cat_row(cat)}, 0)'
        )
        style_cell(ws, info_row, c_offset + 7, pct_formula,
                   font=Font(name="Calibri", size=10, bold=True, color=pal["primary"]),
                   number_format="0%")

    return ws


def _summary_cat_row(cat):
    """Return the row on the Summary sheet where a category's totals live."""
    # Layout: row 3=Income total, then each expense cat gets a row
    # We'll build: 3=Income, 5=Total Budgeted, 7=Total Spent, 9=Remaining
    # Cat totals start row 12 onwards, 2 rows each
    expense_cats = [c for c in CATEGORY_ORDER if c != "Income"]
    if cat == "Income":
        return 3
    idx = expense_cats.index(cat)
    return 12 + idx * 2


# ── 2. MONTHLY LOG SHEET ───────────────────────────────────────────
def build_monthly_log(wb):
    ws = wb.create_sheet("Monthly Log")
    ws.sheet_properties.tabColor = "78909C"

    # Column layout: A=Date, B=Category, C=Item, D=Amount, E=Note
    widths = {"A": 14, "B": 18, "C": 24, "D": 14, "E": 30}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Header row
    ws.merge_cells("A1:E1")
    style_cell(ws, 1, 1, "MONTHLY TRANSACTION LOG",
               font=Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="546E7A"),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 40
    fill_range(ws, 1, 1, 1, 5, PatternFill("solid", fgColor="546E7A"))

    headers = ["Date", "Category", "Item", "Amount (SEK)", "Note"]
    header_fill = PatternFill("solid", fgColor="37474F")
    for i, h in enumerate(headers, 1):
        style_cell(ws, 2, i, h,
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=header_fill,
                   alignment=Alignment(horizontal="center", vertical="center"),
                   border=HEADER_BORDER)
    ws.row_dimensions[2].height = 30

    # Pre-fill example rows + many blank rows for user entry
    examples = [
        ("2026-03-01", "Misc.", "Groceries – ICA", 850, "Weekly shop"),
        ("2026-03-03", "Misc.", "Going out meals – lunch", 145, ""),
        ("2026-03-05", "Subscriptions", "Spotify", 129, "Monthly"),
        ("2026-03-07", "House Loan", "Electricity – Vattenfall", 780, "March bill"),
        ("2026-03-10", "Misc.", "Going out meals – dinner", 320, ""),
        ("2026-03-15", "Savings", "Emergency fund", 2000, "Auto-transfer"),
    ]

    row = 3
    alt_fill_a = PatternFill("solid", fgColor="FAFAFA")
    alt_fill_b = PatternFill("solid", fgColor="FFFFFF")
    for date_str, cat, item, amount, note in examples:
        fill = alt_fill_a if (row % 2 == 1) else alt_fill_b
        style_cell(ws, row, 1, date_str,
                   font=Font(name="Calibri", size=10), fill=fill, border=THIN_BORDER,
                   number_format="YYYY-MM-DD")
        style_cell(ws, row, 2, cat,
                   font=Font(name="Calibri", size=10), fill=fill, border=THIN_BORDER)
        style_cell(ws, row, 3, item,
                   font=Font(name="Calibri", size=10), fill=fill, border=THIN_BORDER)
        style_cell(ws, row, 4, amount,
                   font=Font(name="Calibri", size=10), fill=fill, border=THIN_BORDER,
                   number_format=sek_fmt(""))
        style_cell(ws, row, 5, note,
                   font=Font(name="Calibri", size=10, color="757575"), fill=fill,
                   border=THIN_BORDER)
        row += 1

    # Blank entry rows (500 rows for the user)
    max_entry_row = row + 500
    for r in range(row, max_entry_row):
        fill = alt_fill_a if (r % 2 == 1) else alt_fill_b
        for c in range(1, 6):
            style_cell(ws, r, c, fill=fill, border=THIN_BORDER)
            if c == 4:
                ws.cell(row=r, column=c).number_format = sek_fmt("")
            if c == 1:
                ws.cell(row=r, column=c).number_format = "YYYY-MM-DD"

    # Data validation for Category column (B3:B...)
    from openpyxl.worksheet.datavalidation import DataValidation
    cat_list = ",".join(CATEGORY_ORDER)
    dv = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
    dv.error = "Please select a valid category"
    dv.errorTitle = "Invalid Category"
    dv.prompt = "Choose a budget category"
    dv.promptTitle = "Category"
    ws.add_data_validation(dv)
    dv.add(f"B3:B{max_entry_row}")

    # Data validation for Item column — we'll use a loose list of known items
    all_items = []
    for items in BUDGET_ITEMS.values():
        all_items.extend([name for name, _ in items])

    # Freeze top rows
    ws.freeze_panes = "A3"

    # Auto-filter
    ws.auto_filter.ref = f"A2:E{max_entry_row}"

    return ws, max_entry_row


# ── 3. CATEGORY DETAIL SHEETS ──────────────────────────────────────
def build_category_sheet(wb, cat, log_max_row):
    ws = wb.create_sheet(cat)
    ws.sheet_state = 'hidden'  # Hide category sheets from user view
    pal = COLORS[cat]
    ws.sheet_properties.tabColor = pal["primary"]

    items = BUDGET_ITEMS[cat]

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    # Title
    ws.merge_cells("A1:G1")
    style_cell(ws, 1, 1, cat.upper(),
               font=Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor=pal["primary"]),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 42
    fill_range(ws, 1, 1, 1, 7, PatternFill("solid", fgColor=pal["primary"]))

    # Subtitle
    ws.merge_cells("A2:G2")
    desc = {
        "Income": "All income sources for the month",
        "House Loan": "Housing-related fixed costs",
        "Insurances": "Monthly insurance premiums",
        "Subscriptions": "Recurring subscription services",
        "Savings": "Savings goals & transfers",
        "Misc.": "Variable & discretionary spending",
    }
    style_cell(ws, 2, 1, desc.get(cat, ""),
               font=Font(name="Calibri", size=10, italic=True, color=pal["accent"]),
               fill=PatternFill("solid", fgColor=pal["light"]),
               alignment=Alignment(horizontal="center", vertical="center"))
    fill_range(ws, 2, 2, 1, 7, PatternFill("solid", fgColor=pal["light"]))
    ws.row_dimensions[2].height = 24

    # Column headers
    if cat == "Income":
        col_headers = ["#", "Source", "Budgeted (SEK)", "Actual (SEK)", "", "", ""]
    else:
        col_headers = ["#", "Item", "Budget (SEK)", "Spent from Log (SEK)",
                       "Spent Total (SEK)", "Remaining (SEK)", "% Used"]

    header_fill = PatternFill("solid", fgColor=pal["primary"])
    for i, h in enumerate(col_headers, 1):
        style_cell(ws, 3, i, h,
                   font=Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
                   fill=header_fill,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                   border=HEADER_BORDER)
    ws.row_dimensions[3].height = 32

    # Data rows
    row = 4
    for j, (item_name, budget_val) in enumerate(items):
        alt = PatternFill("solid", fgColor=pal["light"] if j % 2 == 0 else "FFFFFF")

        style_cell(ws, row, 1, j + 1,
                   font=Font(name="Calibri", size=10, color="9E9E9E"), fill=alt,
                   border=THIN_BORDER,
                   alignment=Alignment(horizontal="center", vertical="center"))
        style_cell(ws, row, 2, item_name,
                   font=Font(name="Calibri", size=10, color="212121"), fill=alt,
                   border=THIN_BORDER)
        style_cell(ws, row, 3, budget_val,
                   font=Font(name="Calibri", size=10, color="212121"), fill=alt,
                   border=THIN_BORDER, number_format=sek_fmt(""))

        if cat == "Income":
            # For income, "Actual" = manually entered or SUMIFS from log
            sumifs = (
                f"=SUMPRODUCT(('Monthly Log'!B$3:B${log_max_row}=\"{cat}\")"
                f"*('Monthly Log'!C$3:C${log_max_row}=\"*\"&B{row}&\"*\")"
                f"*'Monthly Log'!D$3:D${log_max_row})"
            )
            style_cell(ws, row, 4, budget_val,  # Default: same as budgeted
                       font=Font(name="Calibri", size=10, color="212121"), fill=alt,
                       border=THIN_BORDER, number_format=sek_fmt(""))
        else:
            # SUMPRODUCT to match category + item substring from log
            # Matches rows where Category=cat AND Item contains the item name
            sumifs = (
                f'=SUMPRODUCT(("Monthly Log"!B$3:B${log_max_row}="{cat}")'
                f'*ISNUMBER(SEARCH(B{row},"Monthly Log"!C$3:C${log_max_row}))'
                f'*"Monthly Log"!D$3:D${log_max_row})'
            )
            # Use simpler SUMIFS approach that works better in Google Sheets
            sumifs_simple = (
                f"=SUMPRODUCT(('Monthly Log'!B$3:B${log_max_row}=\"{cat}\")"
                f"*(ISNUMBER(SEARCH(B{row},'Monthly Log'!C$3:C${log_max_row})))"
                f"*('Monthly Log'!D$3:D${log_max_row}))"
            )
            style_cell(ws, row, 4, sumifs_simple,
                       font=Font(name="Calibri", size=10, color="616161"), fill=alt,
                       border=THIN_BORDER, number_format=sek_fmt(""))

            # E = manual override or same as D (user can add manual entries)
            style_cell(ws, row, 5, f"=D{row}",
                       font=Font(name="Calibri", size=10, color="212121"), fill=alt,
                       border=THIN_BORDER, number_format=sek_fmt(""))

            # F = Remaining
            style_cell(ws, row, 6, f"=C{row}-E{row}",
                       font=Font(name="Calibri", size=10,
                                 color=pal["primary"]),
                       fill=alt, border=THIN_BORDER, number_format=sek_fmt(""))

            # G = % Used
            style_cell(ws, row, 7, f"=IFERROR(E{row}/C{row}, 0)",
                       font=Font(name="Calibri", size=10, bold=True,
                                 color=pal["primary"]),
                       fill=alt, border=THIN_BORDER, number_format="0%")

        row += 1

    # Extra blank rows for user additions
    blank_start = row
    for k in range(10):
        alt = PatternFill("solid", fgColor=pal["light"] if (row - 4) % 2 == 0 else "FFFFFF")
        style_cell(ws, row, 1, "",
                   fill=alt, border=THIN_BORDER)
        for c in range(2, 8):
            style_cell(ws, row, c, "", fill=alt, border=THIN_BORDER)
            if c in (3, 4, 5, 6):
                ws.cell(row=row, column=c).number_format = sek_fmt("")
            if c == 7:
                ws.cell(row=row, column=c).number_format = "0%"
        row += 1
    last_data_row = row - 1

    # Totals row
    total_fill = PatternFill("solid", fgColor=pal["accent"])
    style_cell(ws, row, 1, "",
               fill=total_fill, border=HEADER_BORDER)
    style_cell(ws, row, 2, f"TOTAL — {cat.upper()}",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=total_fill, border=HEADER_BORDER)
    style_cell(ws, row, 3, f"=SUM(C4:C{last_data_row})",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))

    if cat == "Income":
        style_cell(ws, row, 4, f"=SUM(D4:D{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
    else:
        style_cell(ws, row, 4, f"=SUM(D4:D{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
        style_cell(ws, row, 5, f"=SUM(E4:E{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
        style_cell(ws, row, 6, f"=SUM(F4:F{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
        style_cell(ws, row, 7, f"=IFERROR(E{row}/C{row}, 0)",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format="0%")
    fill_range(ws, row, row, 1, 7, total_fill, HEADER_BORDER)
    # Re-apply text (fill_range overwrites)
    style_cell(ws, row, 2, f"TOTAL — {cat.upper()}",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=total_fill, border=HEADER_BORDER)
    style_cell(ws, row, 3, f"=SUM(C4:C{last_data_row})",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
    if cat != "Income":
        style_cell(ws, row, 4, f"=SUM(D4:D{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
        style_cell(ws, row, 5, f"=SUM(E4:E{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
        style_cell(ws, row, 6, f"=C{row}-E{row}",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))
        style_cell(ws, row, 7, f"=IFERROR(E{row}/C{row}, 0)",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format="0%")
    else:
        style_cell(ws, row, 4, f"=SUM(D4:D{last_data_row})",
                   font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
                   fill=total_fill, border=HEADER_BORDER, number_format=sek_fmt(""))

    ws.freeze_panes = "A4"
    return ws, row  # return totals row number


# ── 4. SUMMARY SHEET ───────────────────────────────────────────────
def build_summary(wb, cat_total_rows):
    ws = wb.create_sheet("Summary")
    ws.sheet_state = 'hidden'  # Hide summary helper sheet from user view
    ws.sheet_properties.tabColor = "263238"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    # Title
    ws.merge_cells("A1:F1")
    style_cell(ws, 1, 1, "BUDGET SUMMARY",
               font=Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="263238"),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 44
    fill_range(ws, 1, 1, 1, 6, PatternFill("solid", fgColor="263238"))

    # Row 2: spacer
    fill_range(ws, 2, 2, 1, 6, PatternFill("solid", fgColor="ECEFF1"))

    # Row 3: Total Income
    income_total_row = cat_total_rows["Income"]
    style_cell(ws, 3, 2, "Total Monthly Income",
               font=Font(name="Calibri", size=12, bold=True, color="2E7D32"),
               fill=PatternFill("solid", fgColor="C8E6C9"),
               border=THIN_BORDER)
    style_cell(ws, 3, 3, f"='Income'!D{income_total_row}",
               font=Font(name="Calibri", size=14, bold=True, color="2E7D32"),
               fill=PatternFill("solid", fgColor="C8E6C9"),
               border=THIN_BORDER, number_format=sek_fmt(""))
    fill_range(ws, 3, 3, 1, 6, PatternFill("solid", fgColor="C8E6C9"), THIN_BORDER)
    # Re-apply
    style_cell(ws, 3, 2, "Total Monthly Income",
               font=Font(name="Calibri", size=12, bold=True, color="2E7D32"),
               fill=PatternFill("solid", fgColor="C8E6C9"), border=THIN_BORDER)
    style_cell(ws, 3, 3, f"='Income'!D{income_total_row}",
               font=Font(name="Calibri", size=14, bold=True, color="2E7D32"),
               fill=PatternFill("solid", fgColor="C8E6C9"),
               border=THIN_BORDER, number_format=sek_fmt(""))

    # Row 4: spacer
    fill_range(ws, 4, 4, 1, 6, PatternFill("solid", fgColor="ECEFF1"))

    # Row 5: Total Budgeted Expenses
    expense_cats = [c for c in CATEGORY_ORDER if c != "Income"]
    budget_sum = "+".join([f"'{c}'!C{cat_total_rows[c]}" for c in expense_cats])
    style_cell(ws, 5, 2, "Total Budgeted Expenses",
               font=Font(name="Calibri", size=12, bold=True, color="37474F"),
               fill=PatternFill("solid", fgColor="ECEFF1"), border=THIN_BORDER)
    style_cell(ws, 5, 3, f"={budget_sum}",
               font=Font(name="Calibri", size=14, bold=True, color="37474F"),
               fill=PatternFill("solid", fgColor="ECEFF1"),
               border=THIN_BORDER, number_format=sek_fmt(""))
    fill_range(ws, 5, 5, 1, 6, PatternFill("solid", fgColor="ECEFF1"), THIN_BORDER)
    style_cell(ws, 5, 2, "Total Budgeted Expenses",
               font=Font(name="Calibri", size=12, bold=True, color="37474F"),
               fill=PatternFill("solid", fgColor="ECEFF1"), border=THIN_BORDER)
    style_cell(ws, 5, 3, f"={budget_sum}",
               font=Font(name="Calibri", size=14, bold=True, color="37474F"),
               fill=PatternFill("solid", fgColor="ECEFF1"),
               border=THIN_BORDER, number_format=sek_fmt(""))

    # Row 6: spacer
    fill_range(ws, 6, 6, 1, 6, PatternFill("solid", fgColor="FFFFFF"))

    # Row 7: Total Spent
    spent_sum = "+".join([f"'{c}'!E{cat_total_rows[c]}" for c in expense_cats])
    style_cell(ws, 7, 2, "Total Spent (Actual)",
               font=Font(name="Calibri", size=12, bold=True, color="C62828"),
               fill=PatternFill("solid", fgColor="FFCDD2"), border=THIN_BORDER)
    style_cell(ws, 7, 3, f"={spent_sum}",
               font=Font(name="Calibri", size=14, bold=True, color="C62828"),
               fill=PatternFill("solid", fgColor="FFCDD2"),
               border=THIN_BORDER, number_format=sek_fmt(""))
    fill_range(ws, 7, 7, 1, 6, PatternFill("solid", fgColor="FFCDD2"), THIN_BORDER)
    style_cell(ws, 7, 2, "Total Spent (Actual)",
               font=Font(name="Calibri", size=12, bold=True, color="C62828"),
               fill=PatternFill("solid", fgColor="FFCDD2"), border=THIN_BORDER)
    style_cell(ws, 7, 3, f"={spent_sum}",
               font=Font(name="Calibri", size=14, bold=True, color="C62828"),
               fill=PatternFill("solid", fgColor="FFCDD2"),
               border=THIN_BORDER, number_format=sek_fmt(""))

    # Row 8: spacer
    fill_range(ws, 8, 8, 1, 6, PatternFill("solid", fgColor="FFFFFF"))

    # Row 9: Remaining
    style_cell(ws, 9, 2, "Remaining (Income − Spent)",
               font=Font(name="Calibri", size=12, bold=True, color="1565C0"),
               fill=PatternFill("solid", fgColor="BBDEFB"), border=THIN_BORDER)
    style_cell(ws, 9, 3, "=C3-C7",
               font=Font(name="Calibri", size=14, bold=True, color="1565C0"),
               fill=PatternFill("solid", fgColor="BBDEFB"),
               border=THIN_BORDER, number_format=sek_fmt(""))
    fill_range(ws, 9, 9, 1, 6, PatternFill("solid", fgColor="BBDEFB"), THIN_BORDER)
    style_cell(ws, 9, 2, "Remaining (Income − Spent)",
               font=Font(name="Calibri", size=12, bold=True, color="1565C0"),
               fill=PatternFill("solid", fgColor="BBDEFB"), border=THIN_BORDER)
    style_cell(ws, 9, 3, "=C3-C7",
               font=Font(name="Calibri", size=14, bold=True, color="1565C0"),
               fill=PatternFill("solid", fgColor="BBDEFB"),
               border=THIN_BORDER, number_format=sek_fmt(""))

    # Row 10: spacer
    fill_range(ws, 10, 10, 1, 6, PatternFill("solid", fgColor="FFFFFF"))

    # Row 11: Category breakdown header
    style_cell(ws, 11, 2, "Category",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=HEADER_BORDER)
    style_cell(ws, 11, 3, "Budgeted",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=HEADER_BORDER)
    style_cell(ws, 11, 4, "Spent",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=HEADER_BORDER)
    style_cell(ws, 11, 5, "Remaining",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=HEADER_BORDER)
    style_cell(ws, 11, 6, "% Used",
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="37474F"),
               alignment=Alignment(horizontal="center", vertical="center"),
               border=HEADER_BORDER)
    fill_range(ws, 11, 11, 1, 1, PatternFill("solid", fgColor="37474F"), HEADER_BORDER)
    ws.row_dimensions[11].height = 28

    # Category rows (row 12+, must match _summary_cat_row)
    r = 12
    for cat in expense_cats:
        pal = COLORS[cat]
        tr = cat_total_rows[cat]
        cat_fill = PatternFill("solid", fgColor=pal["light"])

        style_cell(ws, r, 1, "", fill=cat_fill, border=THIN_BORDER)
        style_cell(ws, r, 2, cat,
                   font=Font(name="Calibri", size=11, bold=True, color=pal["primary"]),
                   fill=cat_fill, border=THIN_BORDER)
        style_cell(ws, r, 3, f"='{cat}'!C{tr}",
                   font=Font(name="Calibri", size=11, color="212121"),
                   fill=cat_fill, border=THIN_BORDER, number_format=sek_fmt(""))
        style_cell(ws, r, 4, f"='{cat}'!E{tr}",
                   font=Font(name="Calibri", size=11, color="212121"),
                   fill=cat_fill, border=THIN_BORDER, number_format=sek_fmt(""))
        style_cell(ws, r, 5, f"=C{r}-D{r}",
                   font=Font(name="Calibri", size=11, color=pal["primary"]),
                   fill=cat_fill, border=THIN_BORDER, number_format=sek_fmt(""))
        style_cell(ws, r, 6, f"=IFERROR(D{r}/C{r}, 0)",
                   font=Font(name="Calibri", size=11, bold=True, color=pal["primary"]),
                   fill=cat_fill, border=THIN_BORDER, number_format="0%")
        ws.row_dimensions[r].height = 26

        # Spacer row
        r += 1
        fill_range(ws, r, r, 1, 6, PatternFill("solid", fgColor="FFFFFF"))
        r += 1

    # ── Bar chart: Budget vs Spent per category ──
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Budget vs Actual by Category"
    chart.y_axis.title = "SEK"
    chart.x_axis.title = "Category"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    # We need a small data table for the chart
    chart_data_start = r + 2
    style_cell(ws, chart_data_start, 2, "Category",
               font=Font(name="Calibri", size=9, bold=True))
    style_cell(ws, chart_data_start, 3, "Budgeted",
               font=Font(name="Calibri", size=9, bold=True))
    style_cell(ws, chart_data_start, 4, "Spent",
               font=Font(name="Calibri", size=9, bold=True))
    for i, cat in enumerate(expense_cats):
        cr = chart_data_start + 1 + i
        style_cell(ws, cr, 2, cat)
        # Reference the summary rows
        summary_row = _summary_cat_row(cat)
        style_cell(ws, cr, 3, f"=C{summary_row}", number_format=sek_fmt(""))
        style_cell(ws, cr, 4, f"=D{summary_row}", number_format=sek_fmt(""))

    cats_ref = Reference(ws, min_col=2, min_row=chart_data_start + 1,
                         max_row=chart_data_start + len(expense_cats))
    budgeted_ref = Reference(ws, min_col=3, min_row=chart_data_start,
                             max_row=chart_data_start + len(expense_cats))
    spent_ref = Reference(ws, min_col=4, min_row=chart_data_start,
                          max_row=chart_data_start + len(expense_cats))

    chart.add_data(budgeted_ref, titles_from_data=True)
    chart.add_data(spent_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "90CAF9"
    chart.series[1].graphicalProperties.solidFill = "EF5350"

    ws.add_chart(chart, f"B{chart_data_start + len(expense_cats) + 2}")

    ws.freeze_panes = "A2"
    return ws


# ── 5. INSTRUCTIONS SHEET ──────────────────────────────────────────
def build_instructions(wb):
    ws = wb.create_sheet("How to Use")
    ws.sheet_state = 'hidden'  # Hide instructions from main view
    ws.sheet_properties.tabColor = "FFAB00"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 80

    ws.merge_cells("A1:B1")
    style_cell(ws, 1, 1, "HOW TO USE THIS BUDGET TRACKER",
               font=Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
               fill=PatternFill("solid", fgColor="FF8F00"),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 42
    fill_range(ws, 1, 1, 1, 2, PatternFill("solid", fgColor="FF8F00"))

    instructions = [
        ("1. Record transactions", 
         "Go to the 'Monthly Log' sheet. Enter the date, select a category from the dropdown, "
         "type the item name (include the budget item keyword, e.g. 'Going out meals – lunch'), "
         "enter the amount in SEK, and optionally add a note."),
        ("2. Automatic aggregation",
         "Each category sheet (House Loan, Insurances, etc.) uses SUMPRODUCT formulas to "
         "automatically pull spending from the Monthly Log. The 'Spent from Log' column "
         "searches for the item name in your transaction descriptions."),
        ("3. Dashboard charts",
         "The Dashboard shows donut charts for each category. As you spend, the 'Spent' "
         "segment fills up — like a loading wheel. Green = under budget, filling up = caution."),
        ("4. Summary overview",
         "The Summary sheet shows total income vs total expenses, with a bar chart comparison. "
         "The 'Remaining' value shows how much of your income is unallocated after spending."),
        ("5. Adding new budget items",
         "Each category sheet has 10 blank rows below the pre-filled items. Simply type a new "
         "item name and budget amount. The formulas will extend automatically. Then use the "
         "same item name keyword when logging transactions."),
        ("6. Adjusting budgets",
         "Change any budget amount in column C of any category sheet. All totals and charts "
         "update automatically via formulas."),
        ("7. Income tracking",
         "Update the Income sheet with your actual income received. The 'Actual' column "
         "defaults to the budgeted amount — change it when your actual income differs."),
        ("8. Google Sheets import",
         "Upload this .xlsx file to Google Drive, then open with Google Sheets. "
         "All formulas, charts, and formatting are compatible."),
    ]

    row = 3
    for title, body in instructions:
        style_cell(ws, row, 1, "", fill=PatternFill("solid", fgColor="FFF8E1"))
        style_cell(ws, row, 2, title,
                   font=Font(name="Calibri", size=12, bold=True, color="E65100"),
                   fill=PatternFill("solid", fgColor="FFF8E1"))
        ws.row_dimensions[row].height = 24
        row += 1
        style_cell(ws, row, 1, "", fill=PatternFill("solid", fgColor="FFFFFF"))
        style_cell(ws, row, 2, body,
                   font=Font(name="Calibri", size=10, color="424242"),
                   alignment=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[row].height = 50
        row += 2

    return ws


# ── MAIN BUILD ──────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Build Monthly Log first (needed for row reference)
    log_ws, log_max_row = build_monthly_log(wb)

    # Build each category sheet
    cat_total_rows = {}
    for cat in CATEGORY_ORDER:
        _, total_row = build_category_sheet(wb, cat, log_max_row)
        cat_total_rows[cat] = total_row

    # Build Summary (depends on category total rows)
    build_summary(wb, cat_total_rows)

    # Build Dashboard (depends on Summary being built)
    build_dashboard(wb)

    # Build instructions
    build_instructions(wb)

    # Reorder: Dashboard first, then Monthly Log, categories, Summary, How to Use
    desired_order = (
        ["Dashboard", "Monthly Log"]
        + CATEGORY_ORDER
        + ["Summary", "How to Use"]
    )
    sheet_order = []
    for name in desired_order:
        if name in wb.sheetnames:
            sheet_order.append(wb.sheetnames.index(name))
    wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))

    output_path = "budget_tracker.xlsx"
    wb.save(output_path)
    print(f"Budget tracker saved to: {output_path}")
    print("Upload this file to Google Drive and open with Google Sheets.")


if __name__ == "__main__":
    main()
